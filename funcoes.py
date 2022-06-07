# outras bibliotecas
from os import listdir, getcwd
from io import BytesIO
import openpyxl as xls
import re
import smtplib
from email.mime.base import MIMEBase
from email import encoders


# meus imports
import constantes as const


class ErroCarregamentoExcel(Exception):
    pass
class ErroLoginGmail(Exception):
    pass

# abre o arquivo de texto e retona uma lista com as linhas do documento
def abrir_txt():
    # pegar a loc da pasta de textos e criar uma lista com todos os arquivos
    loc_txt = getcwd() + "\\texto" 
    list_txt = listdir(loc_txt)

    if len(list_txt) > 1: # se ha mais de um arquivo na pasta ele pergunta qual vc quer usar
        
        i = 1
        print("0 sair do programa")
        for texto in list_txt:# mostra para o usuario todas os textos disponiveis
            print(i, texto)
            i = i+1
        texto_escolhido = int(input("Qual destes testos você quer? Digite apenas o numero "))

        if texto_escolhido == 0:
            quit()
        
        if list_txt[texto_escolhido - 1].endswith("txt"): # testa se o arquivo é de texto
            loc_f = loc_txt + "\\" + list_txt[texto_escolhido-1]
        else: 
            raise TypeError("o arquivo que escolheu não termina com .txt =(")

    else:
        if list_txt[0].endswith("txt"): # testa se o arquivo é de texto
            loc_f = loc_txt + "\\" + list_txt[0]
        else: 
            raise Exception("O arquivo que esta na pasta não é um txt =(\nTroque-o e rode o programa de novo")

    with open (loc_f, "r") as arq_txt: # abre o arquivo texto e copia para uma lista
        txt_list = arq_txt.readlines()
        arq_txt.close()
    
    return txt_list

# abre o arquivo de excel, seleciona a planilhs (se for mais de uma pergunta para o usuaria qual usar), retorna a sheet que sera usada. Tambem reajusta o tamanho max da sheet
def abrir_xsl():
    loc_planilhas = getcwd() + "\planilhas"
    list_planilhas = listdir(loc_planilhas)

    if len(list_planilhas) > 1: # se há mais de uma planilha na pasta ele pergunta qual vc quer usar
        
        i = 1
        print("0 sair do programa")
        for plan in list_planilhas:# mostra para o usuario todas as planilhas disponiveis
            print(i, plan)
            i = i+1

        planilha_escolhida = int(input("Qual destas planilha você quer? Digite apenas o numero "))

        if planilha_escolhida == 0:
            quit()

        # teste de erro no input do usuario e determinar loc final da planilha
        if planilha_escolhida.type() !=  int or planilha_escolhida > len(list_planilhas) or planilha_escolhida < 0:
            raise TypeError("Por favor, digite um numero que esteja estre as opções :)")

        elif list_planilhas[planilha_escolhida - 1].endswith("xlsx"):
            loc_final = loc_planilhas + "\\" + list_planilhas[planilha_escolhida-1]

        else: 
            raise TypeError("A planilha que você escolheu não é um arquivo .xlsx =(")

    # se existe apenas 1 opcao
    else:
        if list_planilhas[0].endswith("xlsx"):
            loc_final = loc_planilhas + "\\" + list_planilhas[0]
        else: 
            raise Exception("O arquivo dentro da pasta nao eh uma planilha excel .xlsx =(")

    # abrir a planilha
    try:
        
        # fechar o arquivo seria mandatorio quando usa-se read only, wb.close() nao funcionou, entao usei esse metodo
        with open(loc_final, "rb") as f:
            in_mem_file = BytesIO(f.read())
        # deixo o link do stack overflow onde "Patrick Conwell" deu essa sugestao https://stackoverflow.com/questions/31416842/openpyxl-does-not-close-excel-workbook-in-read-only-mode

        wb = xls.load_workbook(in_mem_file, read_only= True)
        sheet = wb[const.NOME_SHEET]

    except BaseException as err:
        print("para geeks: ", err)
        raise ErroCarregamentoExcel("O carregamento do excel falhou (ㆆ_ㆆ)¯\n")

    # ressetar dimencoes que reedOnly pode ter cagado, para o nosso codigo so importa o numero de linhas
    for k in range (1, sheet.max_row):
        valor = sheet.cell(row= k, column= 1).value
        if valor == None:
            sheet.max_row = k - 1
            break

    return sheet

# usa re para verificar a validade e retorna o gmail_user colocado pelo usuario
def gmail_user_reciver():
    modeloEmailRegex = '^[a-z0-9]+[\._]?[a-z0-9]+[@]\w+[.]\w+$'

    gmail_user = input("De qual email quer enviar?: ")

    if not re.search(modeloEmailRegex, gmail_user): # se o email digitade nao bate com o padrao regex rebutamos a secao e libera outra tentativa
        raise TypeError("o que voce digitou nao eh um email que nos reconhecemos (ㆆ_ㆆ)")

    return gmail_user

# faz o login e retorna o servidor
def gmail_loginer(gmail_user, gmail_password):
    # vamos fazer login na conta gmail e sobreescrever a senha
    try:    
        server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
        server.ehlo()
        server.login(gmail_user, gmail_password)
        gmail_password = "shinanigans_de_proteção"
    except smtplib.SMTPException as err:
        raise ErroLoginGmail("Erro ao fazer login :/\npara geeks: SMTPException- ", err)
    except RuntimeError as err:
        raise ErroLoginGmail("Erro ao fazer login :/\npara geeks: RuntimeError- ", err)
    except SystemExit as err:
        raise ErroLoginGmail("Erro ao fazer login :/\npara geeks: SystemExit- ", err)
    except:
        raise ErroLoginGmail("Erro ao fazer login :/")
    
    return server


# retorna o texto com os @@@ subtituidos pelos termos no excel
def construtor_texto(txt_list, linha, sheet):
    # Passo 1: transformar a lista em uma matriz
    txt_matriz = [[None] for k in range(len(txt_list))] # matriz com n linhas n = len(txt_list)
    for i in range(len(txt_list)):
        txt_matriz[i] = txt_list[i].split() # cada linha da matriz eh quebrada por palavra

    coluna = const.COL_GAPS

    # Passo 2: substituir @@@
    for n in range(len(txt_matriz)):
        for m in range(len(txt_matriz[n])):
            if txt_matriz[n][m] == "@@@":
                txt_matriz[n][m] = sheet.cell(row= linha, column= coluna).value
                coluna = coluna + 1

    # Passo 3: recontruir a string
    texto = ""
    for n in range(len(txt_matriz)):
        for m in range(len(txt_matriz[n])):
            texto = texto + txt_matriz[n][m] + " "
        texto = texto + "\n"

    return texto

def anexar(loc, list_permitidos, arq, msg):
    #testar se o arq esta nos permitidos
    boleano = False
    for terminacao in list_permitidos:
        if loc.endswith(terminacao):
            boleano = True
    
    if boleano:
        # cria o obj mime que vai pra msg como anexo
        with open (loc, 'rb') as f: 
            data = MIMEBase('application', 'octet-stream')
            data.set_payload(f.read())
            f.close() 
        encoders.encode_base64(data)
        data.add_header('Content-Disposition', 'attachment', filename= arq)
        msg.attach(data)          
    
    else:
        print("não reconhecemos arquivos do tipo '%s' ou não podemos envia-lo" %arq)


    